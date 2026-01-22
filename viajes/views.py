from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Vehiculo
from .forms import LoginForm


def login_view(request):
    if request.user.is_authenticated:
        return redirect('viajes:lista')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Bienvenido, {user.username}')
            return redirect('viajes:lista')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    else:
        form = LoginForm()
    
    return render(request, 'registration/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión correctamente')
    return redirect('viajes:login')


@login_required
def lista_viajes(request):
    vehiculos = Vehiculo.objects.all()
    
    # Filtros
    placa = request.GET.get('placa', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado_validacion = request.GET.get('estado_validacion', '')
    
    # Aplicar filtro de placa
    if placa:
        vehiculos = vehiculos.filter(placa__icontains=placa)
    
    # Aplicar filtro de rango de fechas
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            vehiculos = vehiculos.filter(fecha_inicio__date__gte=fecha_desde_obj.date())
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            vehiculos = vehiculos.filter(fecha_inicio__date__lte=fecha_hasta_obj.date())
        except ValueError:
            pass
    
    # Aplicar filtro de estado de validación
    if estado_validacion == 'validado':
        vehiculos = vehiculos.filter(validado=True)
    elif estado_validacion == 'no_validado':
        vehiculos = vehiculos.filter(validado=False)
    
    # Paginación
    paginator = Paginator(vehiculos, 20)  # 20 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'placa': placa,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'estado_validacion': estado_validacion,
    }
    
    return render(request, 'viajes/lista.html', context)


@login_required
@require_http_methods(["POST"])
def toggle_validado(request, vehiculo_id):
    try:
        vehiculo = Vehiculo.objects.get(id=vehiculo_id)
        vehiculo.validado = not vehiculo.validado
        vehiculo.save()
        return JsonResponse({
            'success': True,
            'validado': vehiculo.validado
        })
    except Vehiculo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vehículo no encontrado'
        }, status=404)


@login_required
def exportar_datos(request):
    vehiculos = Vehiculo.objects.all()
    
    # Aplicar los mismos filtros que en la vista de lista
    placa = request.GET.get('placa', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado_validacion = request.GET.get('estado_validacion', '')
    formato = request.GET.get('formato', 'csv')
    
    if placa:
        vehiculos = vehiculos.filter(placa__icontains=placa)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            vehiculos = vehiculos.filter(fecha_inicio__date__gte=fecha_desde_obj.date())
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            vehiculos = vehiculos.filter(fecha_inicio__date__lte=fecha_hasta_obj.date())
        except ValueError:
            pass
    
    if estado_validacion == 'validado':
        vehiculos = vehiculos.filter(validado=True)
    elif estado_validacion == 'no_validado':
        vehiculos = vehiculos.filter(validado=False)
    
    # Ordenar por fecha_inicio descendente
    vehiculos = vehiculos.order_by('-fecha_inicio')
    
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    
    if formato == 'excel':
        # Exportar a Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehículos"
        
        # Encabezados con estilo
        headers = ['Código', 'Placa', 'Tipo Vehículo', 'Fecha Inicio', 'Fecha Fin', 
                  'Número Entregas', 'Facturación', 'Cliente', 'Validado']
        
        header_fill = PatternFill(start_color="FF7900", end_color="FF7900", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_num, vehiculo in enumerate(vehiculos, 2):
            ws.cell(row=row_num, column=1, value=vehiculo.codigo)
            ws.cell(row=row_num, column=2, value=vehiculo.placa)
            ws.cell(row=row_num, column=3, value=vehiculo.get_tipo_vehiculo_display())
            ws.cell(row=row_num, column=4, value=vehiculo.fecha_inicio.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=5, value=vehiculo.fecha_fin.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=6, value=vehiculo.numero_entregas)
            ws.cell(row=row_num, column=7, value=float(vehiculo.facturacion))
            ws.cell(row=row_num, column=8, value=vehiculo.cliente)
            ws.cell(row=row_num, column=9, value='Sí' if vehiculo.validado else 'No')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="vehiculos_{timestamp}.xlsx"'
        wb.save(response)
        return response
    
    else:
        # Exportar a CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="vehiculos_{timestamp}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Placa', 'Tipo Vehículo', 'Fecha Inicio', 'Fecha Fin', 
                        'Número Entregas', 'Facturación', 'Cliente', 'Validado'])
        
        for vehiculo in vehiculos:
            writer.writerow([
                vehiculo.codigo,
                vehiculo.placa,
                vehiculo.get_tipo_vehiculo_display(),
                vehiculo.fecha_inicio.strftime('%Y-%m-%d %H:%M'),
                vehiculo.fecha_fin.strftime('%Y-%m-%d %H:%M'),
                vehiculo.numero_entregas,
                str(vehiculo.facturacion),
                vehiculo.cliente,
                'Sí' if vehiculo.validado else 'No'
            ])
        
        return response
